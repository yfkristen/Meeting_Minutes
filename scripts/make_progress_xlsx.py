#!/usr/bin/env python3
"""從會議記錄 Markdown 的「各案進度」章節產出個案進度 Excel。

用法:
    python3 scripts/make_progress_xlsx.py minutes/2026-07-23_雙週會.md
    python3 scripts/make_progress_xlsx.py minutes/2026-07-23_雙週會.md -o output/excel/自訂.xlsx

輸出預設為 output/excel/<原檔名>_個案進度.xlsx
※ 欄位為暫定版本，待提供實際 Excel 範本後再調整。
"""

import argparse
import re
import sys
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

HEADERS = ["編號", "個案名稱", "進度說明", "負責人", "狀態", "期限", "備註"]
COL_WIDTHS = [6, 24, 70, 12, 10, 14, 20]

SECTION2 = re.compile(r"^[一二三四五六七八九十]+、")
SECTION1 = re.compile(r"^[壹貳參肆伍陸柒捌玖拾]+、")
ITEM = re.compile(r"^(\d+)\.\s*(.+)$")
OWNER = re.compile(r"[（(]([^（）()]+)[）)]\s*$")

HEAD_FILL = PatternFill("solid", fgColor="1F3864")
HEAD_FONT = Font(name="微軟正黑體", size=11, bold=True, color="FFFFFF")
BODY_FONT = Font(name="微軟正黑體", size=11)
THIN = Side(style="thin", color="BFBFBF")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)


def parse_meeting_title(md_path: Path, lines: list[str]) -> tuple[str, str]:
    """回傳 (會議日期, 會議名稱)，取自檔名 YYYY-MM-DD_名稱。"""
    stem = md_path.stem
    date, _, name = stem.partition("_")
    for line in lines:
        m = re.search(r"「(.+?)」", line)
        if m:
            name = m.group(1)
            break
    return date, name


def parse_items(lines: list[str]) -> list[dict]:
    """擷取「各案進度」章節底下的編號項目。"""
    items, in_section = [], False
    for raw in lines:
        line = raw.strip()
        if not line:
            continue
        if SECTION1.match(line):
            in_section = False
            continue
        if SECTION2.match(line):
            in_section = "各案進度" in line
            continue
        if not in_section:
            continue
        m = ITEM.match(line)
        if not m:
            continue
        no, text = m.group(1), m.group(2)

        owner = ""
        om = OWNER.search(text)
        if om:
            owner = om.group(1).strip()
            text = text[: om.start()].rstrip()

        name, sep, detail = text.partition("：")
        if not sep:
            name, detail = text[:20], text
        items.append({
            "no": no,
            "name": name.strip(),
            "detail": detail.strip(),
            "owner": owner,
        })
    return items


def build(md_path: Path, out_path: Path) -> tuple[Path, int]:
    lines = md_path.read_text(encoding="utf-8").splitlines()
    date, meeting = parse_meeting_title(md_path, lines)
    items = parse_items(lines)

    wb = Workbook()
    ws = wb.active
    ws.title = "個案進度"

    ws["A1"] = f"{meeting}　個案進度追蹤表"
    ws["A1"].font = Font(name="微軟正黑體", size=14, bold=True)
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(HEADERS))
    ws["A2"] = f"會議日期：{date}"
    ws["A2"].font = BODY_FONT
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=len(HEADERS))

    header_row = 4
    for col, title in enumerate(HEADERS, start=1):
        c = ws.cell(row=header_row, column=col, value=title)
        c.fill, c.font, c.border = HEAD_FILL, HEAD_FONT, BORDER
        c.alignment = Alignment(horizontal="center", vertical="center")

    for i, item in enumerate(items):
        row = header_row + 1 + i
        values = [item["no"], item["name"], item["detail"], item["owner"], "進行中", "", ""]
        for col, value in enumerate(values, start=1):
            c = ws.cell(row=row, column=col, value=value)
            c.font, c.border = BODY_FONT, BORDER
            c.alignment = Alignment(
                horizontal="left" if col in (2, 3, 7) else "center",
                vertical="top",
                wrap_text=True,
            )
        ws.row_dimensions[row].height = 60

    for col, width in enumerate(COL_WIDTHS, start=1):
        ws.column_dimensions[get_column_letter(col)].width = width
    ws.freeze_panes = ws.cell(row=header_row + 1, column=1)
    if items:
        ws.auto_filter.ref = (
            f"A{header_row}:{get_column_letter(len(HEADERS))}{header_row + len(items)}"
        )

    out_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(out_path)
    return out_path, len(items)


def main() -> int:
    ap = argparse.ArgumentParser(description="會議記錄各案進度 → Excel")
    ap.add_argument("markdown", help="minutes/ 底下的會議記錄 .md")
    ap.add_argument("-o", "--output", help="輸出 .xlsx 路徑")
    args = ap.parse_args()

    md_path = Path(args.markdown)
    if not md_path.is_file():
        print(f"找不到檔案：{md_path}", file=sys.stderr)
        return 1

    out_path = Path(args.output) if args.output else \
        Path("output/excel") / f"{md_path.stem}_個案進度.xlsx"

    path, count = build(md_path, out_path)
    print(f"已產出：{path}（{count} 筆個案）")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
