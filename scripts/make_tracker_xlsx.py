#!/usr/bin/env python3
"""由 tracker/task_progress.csv 產出跨會議累積的任務追蹤表 Excel。

用法:
    python3 scripts/make_tracker_xlsx.py
    python3 scripts/make_tracker_xlsx.py -i tracker/task_progress.csv -o output/excel/追蹤表.xlsx

資料格式（長表，一列＝某任務在某次會議的進度）：
    編號,任務名稱,負責人,會議日期,狀態,本次進度

產出（寬表，依 skill「ai-pmo-meeting-minutes」的追蹤表規則）：
    一列一任務、各次會議進度並排，該次未討論者自動填「本次未討論」。
    任務名稱、負責人、狀態取該任務最近一次會議的值。
"""

import argparse
import csv
import sys
from collections import OrderedDict
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

FIXED_HEADERS = ["編號", "任務名稱", "負責人", "最新狀態"]
FIXED_WIDTHS = [6, 26, 12, 12]
MEETING_WIDTH = 46
NOT_DISCUSSED = "本次未討論"

TITLE = "金控AI推動委員會PMO　任務追蹤表"
HEAD_FILL = PatternFill("solid", fgColor="1F3864")
HEAD_FONT = Font(name="微軟正黑體", size=11, bold=True, color="FFFFFF")
FIXED_FILL = PatternFill("solid", fgColor="F2F2F2")
BODY_FONT = Font(name="微軟正黑體", size=11)
MUTED_FONT = Font(name="微軟正黑體", size=11, color="A6A6A6", italic=True)
THIN = Side(style="thin", color="BFBFBF")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

STATUS_FILL = {
    "已完成": PatternFill("solid", fgColor="E2EFDA"),
    "進行中": PatternFill("solid", fgColor="FFF2CC"),
    "籌備中": PatternFill("solid", fgColor="DDEBF7"),
    "暫緩": PatternFill("solid", fgColor="FCE4EC"),
}


def task_sort_key(no: str):
    """編號優先數字排序，非數字者排在後面並依字串排。"""
    try:
        return (0, int(no), "")
    except ValueError:
        return (1, 0, no)


def read_rows(csv_path: Path) -> list[dict]:
    with csv_path.open(encoding="utf-8-sig", newline="") as f:
        rows = [
            {(k or "").strip(): (v or "").strip() for k, v in row.items()}
            for row in csv.DictReader(f)
        ]
    required = {"編號", "任務名稱", "會議日期", "本次進度"}
    if rows and not required.issubset(rows[0]):
        missing = "、".join(sorted(required - set(rows[0])))
        raise SystemExit(f"{csv_path} 缺少欄位：{missing}")
    return [r for r in rows if r.get("編號")]


def build(csv_path: Path, out_path: Path) -> tuple[Path, int, int]:
    rows = read_rows(csv_path)
    meetings = sorted({r["會議日期"] for r in rows})

    tasks: "OrderedDict[str, dict]" = OrderedDict()
    for r in sorted(rows, key=lambda r: r["會議日期"]):   # 後出現者覆蓋，取最新值
        t = tasks.setdefault(r["編號"], {"progress": {}})
        t["name"] = r["任務名稱"] or t.get("name", "")
        t["owner"] = r.get("負責人") or t.get("owner", "")
        t["status"] = r.get("狀態") or t.get("status", "")
        t["progress"][r["會議日期"]] = r["本次進度"]

    ordered = sorted(tasks.items(), key=lambda kv: task_sort_key(kv[0]))
    headers = FIXED_HEADERS + meetings

    wb = Workbook()
    ws = wb.active
    ws.title = "任務追蹤"

    ws["A1"] = TITLE
    ws["A1"].font = Font(name="微軟正黑體", size=14, bold=True)
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(headers))
    ws["A2"] = f"更新至：{meetings[-1] if meetings else '—'}　｜　共 {len(ordered)} 項任務"
    ws["A2"].font = Font(name="微軟正黑體", size=10, color="7F7F7F")
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=len(headers))

    head_row = 4
    for col, title in enumerate(headers, start=1):
        c = ws.cell(row=head_row, column=col, value=title)
        c.fill, c.font, c.border = HEAD_FILL, HEAD_FONT, BORDER
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    for i, (no, task) in enumerate(ordered):
        row = head_row + 1 + i
        fixed = [no, task["name"], task["owner"], task["status"]]
        for col, value in enumerate(fixed, start=1):
            c = ws.cell(row=row, column=col, value=value)
            c.font, c.border = BODY_FONT, BORDER
            c.alignment = Alignment(
                horizontal="left" if col == 2 else "center",
                vertical="center", wrap_text=True)
            c.fill = STATUS_FILL.get(value, FIXED_FILL) if col == 4 else FIXED_FILL

        first_seen = min(task["progress"])
        for j, meeting in enumerate(meetings):
            c = ws.cell(row=row, column=len(FIXED_HEADERS) + 1 + j)
            progress = task["progress"].get(meeting)
            if progress:
                c.value, c.font = progress, BODY_FONT
            elif meeting < first_seen:          # 該任務當時尚未成案，留白
                c.value, c.font = "", BODY_FONT
            else:
                c.value, c.font = NOT_DISCUSSED, MUTED_FONT
            c.border = BORDER
            c.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
        ws.row_dimensions[row].height = 62

    for col, width in enumerate(FIXED_WIDTHS, start=1):
        ws.column_dimensions[get_column_letter(col)].width = width
    for j in range(len(meetings)):
        ws.column_dimensions[
            get_column_letter(len(FIXED_HEADERS) + 1 + j)].width = MEETING_WIDTH

    ws.freeze_panes = ws.cell(row=head_row + 1, column=len(FIXED_HEADERS) + 1)
    if ordered:
        ws.auto_filter.ref = (
            f"A{head_row}:{get_column_letter(len(headers))}{head_row + len(ordered)}")

    out_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(out_path)
    return out_path, len(ordered), len(meetings)


def main() -> int:
    ap = argparse.ArgumentParser(description="PMO 任務追蹤表 CSV → Excel")
    ap.add_argument("-i", "--input", default="tracker/task_progress.csv",
                    help="長表 CSV（預設 tracker/task_progress.csv）")
    ap.add_argument("-o", "--output", default="output/excel/PMO任務追蹤表.xlsx",
                    help="輸出 .xlsx 路徑")
    args = ap.parse_args()

    csv_path = Path(args.input)
    if not csv_path.is_file():
        print(f"找不到檔案：{csv_path}", file=sys.stderr)
        return 1

    path, tasks, meetings = build(csv_path, Path(args.output))
    print(f"已產出：{path}（{tasks} 項任務 × {meetings} 次會議）")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
